#단체메일 보내기
from openpyxl import Workbook
import datetime

wb = Workbook()
ws = wb.active

ws['A1'] = 'jjimini98@naver.com'
ws['A2'] = 'lovers2019@kakao.com'

wb.save('Final Exam\\Lecture 11\email_list.xlsx')


